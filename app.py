import sys
from PyQt6.QtWidgets import (
    QMessageBox, QDateEdit, QApplication, QMainWindow, QWidget, QVBoxLayout,
    QHBoxLayout, QLabel, QLineEdit, QPushButton, QFormLayout, QRadioButton,
    QButtonGroup, QTextEdit, QTableWidget, QTableWidgetItem, QComboBox, QInputDialog, QSlider, QTabWidget
)
from PyQt6.QtCore import (
    QDateTime, QDate, Qt
    )
from PyQt6 import QtCore
import tkinter as tk
import tkinter as Tk
from tkinter import (
    simpledialog, messagebox, ttk                 
                     )
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import PieChart, Reference
from datetime import datetime
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
import email
import email.utils
import json
from apscheduler.schedulers.background import BackgroundScheduler
import schedule
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from PyQt6.QtCore import QDate
import mariadb
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter



# Load configuration settings from config.json
with open("config.json", "r") as config_file:
    config = json.load(config_file)

# Get email settings from the config
smtp_server = config["email"]["smtp_server"]
smtp_port = config["email"]["smtp_port"]
sender_email = config["email"]["sender_email"]
sender_password = config["email"]["sender_password"]
recipient_email = config["email"]["recipient_email"]

# Get login settings from the config
host = config["db"]["host"]
user = config["db"]["user"]
password = config["db"]["password"]
database = config["db"]["database"]

#tzs is tzsh

categorieslist = [
            "Utilities",
            "Transportation",
            "Office Supplies",
            "Entertainment",
            "Travel",
            "Food & Drinks",
            "Healthcare",
            "Miscellaneous (with Folio)",
            "Charity and Donations",
            "Advertising",
            "Rent or Mortgage",
            "Repairs and Maintenance",
            "Personal Withdrawings",
            "Wages/Salaries"
            ]

# Database connection parameters, change the username and password to a newly created user
db_config = {
    "host": host,
    "user": user,
    "password": password,
    "database": database
}

# Establish a connection to the MariaDB database
conn = mariadb.connect(**db_config)

# Create a cursor
cursor = conn.cursor()

# Close the cursor and connection when done
cursor.close()
conn.close()



class PettyCashApp(QMainWindow):
    def __init__(self):
        super().__init__()

        # Initialize UI components
        self.setWindowTitle("Petty Cash App")
        self.setGeometry(100, 100, 600, 600)

        # Set the application-wide style sheet
        self.setStyleSheet(
            "QWidget { background-color: #B0C4DE; color: black; }" + #or #F5F5F5
            "QLabel { font-size: 18px;}"
        )

       # Create database connection parameters
        db_config = {
            "host": host,
            "user": user,
            "password": password,
            "database": database
        }
            
        # Establish a connection to the MariaDB database
        self.conn = mariadb.connect(**db_config)
        self.cursor = self.conn.cursor()
        
        self.page_number = 1  # Initialize page_number attribute
        self.rows_per_page = 10
        self.setup_login_page()

    def setup_login_page(self):
        self.central_widget = QWidget(self)
        self.setCentralWidget(self.central_widget)

        layout = QVBoxLayout()
        self.central_widget.setLayout(layout)

        self.username_label = QLabel("Username:")
        self.username_entry = QLineEdit()
        self.password_label = QLabel("Password:")
        self.password_entry = QLineEdit()
        self.password_entry.setEchoMode(QLineEdit.EchoMode.Password)
        self.login_button = QPushButton("Login")
        self.login_button.clicked.connect(self.login)

        layout.addWidget(self.username_label)
        layout.addWidget(self.username_entry)
        layout.addWidget(self.password_label)
        layout.addWidget(self.password_entry)
        layout.addWidget(self.login_button)

    def login(self):
        username = self.username_entry.text()
        password = self.password_entry.text()

        # Check user credentials in the database
        self.cursor.execute("SELECT * FROM users WHERE username = ? AND password = ?", (username, password))
        user = self.cursor.fetchone()

        if user is not None:
                if user[3] == 'y':
                   self.setup_tabbed_interface_admin()
                else:
                   self.setup_tabbed_interface()
                    
        else:
            QMessageBox.critical(self, "Login Error", "Invalid username or password")

    def setup_tabbed_interface(self):
        self.tab_widget = QTabWidget()
        self.setCentralWidget(self.tab_widget)

        self.input_tab = QWidget()
        self.view_tab = QWidget()

        self.setup_input_page()  # Call setup_input_page first to define self.table
        self.setup_view_page()   # Call setup_view_page after self.table is defined
        self.setup_filter_page()
        
        self.tab_widget.setCurrentIndex(0)  # Set the default tab to the input page

    def setup_tabbed_interface_admin(self):
        self.tab_widget = QTabWidget()
        self.setCentralWidget(self.tab_widget)

        self.input_tab = QWidget()
        self.view_tab = QWidget()
        self.search_tab = QWidget()
        self.approve_page = QWidget()

        self.setup_input_page()  # Call setup functions for each tab
        self.setup_view_page()
        self.setup_filter_page()
        self.setup_approve_page() #approve table
        
        self.tab_widget.setCurrentIndex(0)  # Set the default tab to the input page

    def setup_input_page(self):
        self.input_tab = QWidget()
        self.tab_widget.addTab(self.input_tab, "Input Page")

        layout = QVBoxLayout(self.input_tab)

        self.table = QTableWidget()
        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels(["Voucher", "Date", "Description", "Category", "Amount", "Currency", "Folio"])
            
        self.date_label = QLabel("Date:")
        self.date_edit = QDateEdit()
        self.date_edit.setDateTime(QDateTime.currentDateTime())  
        self.date_edit.setCalendarPopup(True)
        self.date_edit != None
        
        self.description_label = QLabel("Description:")
        self.description_entry = QLineEdit()
        self.description_entry != None
        
        self.amount_label = QLabel("Amount:")
        self.amount_entry = QLineEdit()
        self.amount_entry != None

        self.currency_group = QButtonGroup()
        self.tzs_radio = QRadioButton("TZS")
        self.usd_radio = QRadioButton("USD")
        self.currency_group.addButton(self.tzs_radio)
        self.currency_group.addButton(self.usd_radio)

        self.category_label = QLabel("Category:")
        self.category_layout = QFormLayout()
        self.category_group = QButtonGroup()

        self.category_label = QLabel("Category:")
        self.category_layout = QVBoxLayout()  # Use QVBoxLayout to stack the radio buttons

        self.category_group = QButtonGroup()

        # Create two rows of radio buttons for custom categories
        category_radio_buttons = []
        row_layout = QHBoxLayout()
        
        for i, category in enumerate(categorieslist):
            category_radio_button = QRadioButton(category)
            category_radio_buttons.append(category_radio_button)
            self.category_group.addButton(category_radio_button)
            
            # Add the radio button to the current row layout
            row_layout.addWidget(category_radio_button)
            
            # Create a new row layout after every 7 categories (adjust as needed)
            if (i + 1) % 7 == 0 or i == len(categorieslist) - 1:
                self.category_layout.addLayout(row_layout)
                row_layout = QHBoxLayout()

        self.folio_label = QLabel("Folio (optional):")
        self.folio_entry = QLineEdit()

        self.save_button = QPushButton("Save")
        self.save_button.clicked.connect(self.save_data)

        layout.addWidget(self.date_label)
        layout.addWidget(self.date_edit)
        layout.addWidget(self.description_label)
        layout.addWidget(self.description_entry)
        layout.addWidget(self.amount_label)
        layout.addWidget(self.amount_entry)
        layout.addWidget(self.tzs_radio)
        layout.addWidget(self.usd_radio)
        layout.addWidget(self.category_label)
        layout.addLayout(self.category_layout)
        layout.addWidget(self.folio_label)
        layout.addWidget(self.folio_entry)
        layout.addWidget(self.save_button)
        
        self.populate_table()

    def open_calendar_popup(self):
            self.date_edit.setCalendarPopup(True)

    def save_data(self):
        date = self.date_edit.date().toString("yyyy-MM-dd")
        description = self.description_entry.text()
        amount = self.amount_entry.text()
        currency = "USD" if self.usd_radio.isChecked() else "TZS"

        try:
            # Fetch the count of entries for the current date
            today = datetime.now().date()
            self.cursor.execute("SELECT COUNT(*) FROM expenses WHERE date = ?", (today,))
            count_today = self.cursor.fetchone()[0] if self.cursor.rowcount > 0 else 0

            # Calculate the number of digits needed for entry count
            num_digits = len(str(count_today + 1))
            max_digits = len(str(9999))  # Adjust as needed

            # Generate the voucher number based on the count for the current date
            entry_number = count_today + 1

            # Pad entry number with leading zeros to ensure the required number of digits
            entry_number_str = str(entry_number).zfill(num_digits if num_digits < max_digits else max_digits)

            # Check for empty fields and ensure a category is selected
            if not description:
                QMessageBox.critical(self, "Error", "Please fill in the Description.")
                return
            if not amount:
                QMessageBox.critical(self, "Error", "Please fill in the Amount.")
                return
            try:
                float(amount)
            except ValueError:
                QMessageBox.critical(self, "Error", "Amount must be a valid number.")
                return

            # Check if a category is selected
            if not any(button.isChecked() for button in self.category_group.buttons()):
                QMessageBox.critical(self, "Error", "Please select a Category.")
                return

            # Get the selected category
            selected_category = ""
            for button in self.category_group.buttons():
                if button.isChecked():
                    selected_category = button.text()

            # Check if the selected category is "Miscellaneous"
            if selected_category == "Miscellaneous (with Folio)":
                folio_text = self.folio_entry.text()
                if not folio_text.strip():  # Check if folio field is blank or contains only whitespace
                    # Prompt the user for folio input
                    folio, ok = QInputDialog.getText(self, "Folio Input", "Enter Folio:")
                    if ok:
                        folio = folio or "Not Provided"
                    else:
                        # User canceled, do not proceed with saving
                        return
                else:
                    folio = folio_text  # Take folio from the field if not blank
            else:
                # For other categories, set folio to the value in the folio field
                folio = self.folio_entry.text()

            # Insert data into the database with the new voucher number
            voucher_number = f"{today.strftime('%Y-%m-%d')}-{entry_number_str}"
            self.cursor.execute(
                "INSERT INTO expenses (voucher, date, description, category, amount, currency, folio, approved) VALUES (?, ?, ?, ?, ?, ?, ?, 'none')",
                (voucher_number, date, description, selected_category, amount, currency, folio)
            )
            self.conn.commit()
            QMessageBox.information(self, "Success", f"Data saved successfully with Voucher #{voucher_number}")

            # Refresh the "Approve Page" table with the latest data
            self.populate_approve_page(self.approve_table)
        except Exception as e:
            self.conn.rollback()
            QMessageBox.critical(self, "Error", "An error occurred while saving the data: " + str(e))

        # Clear input fields
        self.date_edit.setDate(QDateTime.currentDateTime().date())
        self.description_entry.clear()
        self.amount_entry.clear()
        self.folio_entry.clear()

        self.populate_table()
    
    def setup_view_page(self):
        self.view_tab = QWidget()
        self.tab_widget.addTab(self.view_tab, "View Page")

        layout = QVBoxLayout(self.view_tab)
        
        
        self.table = QTableWidget()
        self.table.setColumnCount(7)
        self.table.setHorizontalHeaderLabels(["              Voucher               ", "             Date               ", "             Description               ", "              Category                ", "                 Amount              ", " Currency ", "                 Folio                "])
        
        self.pagination_layout = QHBoxLayout()
        self.prev_button = QPushButton("Previous")
        self.prev_button.clicked.connect(self.prev_page)
        self.next_button = QPushButton("Next")
        self.next_button.clicked.connect(self.next_page)
        self.pagination_layout.addWidget(self.prev_button)
        self.pagination_layout.addWidget(self.next_button)

        self.edit_button = QPushButton("Edit")
        self.edit_button.clicked.connect(self.edit_entry)
        self.delete_button = QPushButton("Delete")
        self.delete_button.clicked.connect(self.delete_entry)
        self.backup_button = QPushButton("Backup data")  
        self.backup_button.clicked.connect(self.backup_data)
        self.send_email_button = QPushButton("Send Email")
        self.send_email_button.clicked.connect(self.send_email)


        layout.addWidget(self.table)
        layout.addWidget(self.send_email_button)
        layout.addLayout(self.pagination_layout)
        layout.addWidget(self.edit_button)
        layout.addWidget(self.delete_button)  
        layout.addWidget(self.backup_button)    

        self.page_number = 1
        self.rows_per_page = 10

        self.populate_table()

    def backup_data(self):
        try:
            # Fetch data based on the specified date range
            self.cursor.execute("SELECT * FROM expenses ORDER BY id DESC")
            data = self.cursor.fetchall()
            
            if data:
                column_names = [description[0] for description in self.cursor.description]
                df = pd.DataFrame(data, columns=column_names)

                workbook = Workbook()
                worksheet = workbook.active
                worksheet.title = "PettyCash Backup"

                for r_idx, row in enumerate(data, start=2):
                    for c_idx, value in enumerate(row, start=1):
                        if column_names[c_idx - 1] == 'date':
                            # Format the date as dd MM yyyy
                            worksheet.cell(row=r_idx, column=c_idx).value = value.strftime("%d %m %Y")
                        else:
                            worksheet.cell(row=r_idx, column=c_idx).value = value


            # Adjust column widths for all columns
                for col in range(1, len(column_names) + 1):
                    column_letter = get_column_letter(col)
                    worksheet.column_dimensions[column_letter].width = 30  # Set the desired width for all columns


                category_totals = {}
                for row in data:
                    category = row[column_names.index("category")]
                    amount = row[column_names.index("amount")]
                    if category in category_totals:
                        category_totals[category] += amount
                    else:
                        category_totals[category] = amount

                for col, category in enumerate(category_totals.keys(), start=len(column_names) + 2):
                    worksheet.cell(row=1, column=col, value=category)
                    for row, row_data in enumerate(data, start=2):
                        if row_data[column_names.index("category")] == category:
                            worksheet.cell(row=row, column=col, value=row_data[column_names.index("amount")])

                pie_chart = PieChart()
                labels = Reference(worksheet, min_col=len(column_names) + 2, min_row=2, max_row=len(data) + 1, max_col=len(column_names) + len(category_totals) + 1)
                data = Reference(worksheet, min_col=len(column_names) + 1, min_row=2, max_row=len(data) + 1, max_col=len(column_names) + 1)
                pie_chart.add_data(data, titles_from_data=True)
                pie_chart.set_categories(labels)
                worksheet.add_chart(pie_chart, "A10")

                today_date = pd.Timestamp("today").strftime("%Y-%m-%d")
                backup_filename = f"{today_date} PettyCash Backup.xlsx"
                workbook.save(backup_filename)

                QMessageBox.information(self, "Note", f"Data backed up as \'{backup_filename}\'")

            else:
                QMessageBox.information(self, "Note", 'No data to back up')

        except Exception as e:
            return

    
    def sort_table(self, column):
        self.table.sortItems(column)

        # Connect the sectionClicked signal to the sort_table function
        self.table.horizontalHeader().sectionClicked.connect(self.sort_table)

    def setup_approve_page(self):
        self.approve_page = QWidget()
        self.tab_widget.addTab(self.approve_page, "Approve Page")

        layout = QVBoxLayout(self.approve_page)

        self.approve_table = QTableWidget()
        self.approve_table.setColumnCount(7)
        self.approve_table.setHorizontalHeaderLabels(["              Voucher               ", "             Date               ", "             Description               ", "              Category                ", "                 Amount              ", " Currency ", "                 Folio                "])

        approve_layout = QHBoxLayout()
        approve_button = QPushButton("Approve")
        approve_button.clicked.connect(self.approve_entry)
        decline_button = QPushButton("Decline")
        decline_button.clicked.connect(self.decline_entry)
        approve_layout.addWidget(approve_button)
        approve_layout.addWidget(decline_button)

        layout.addWidget(self.approve_table)
        layout.addLayout(approve_layout)

        self.populate_approve_page(self.approve_table)  # Use the correct table widget name

    def populate_approve_page(self, approve_page):
        try:
            self.cursor.execute("SELECT voucher, date, description, category, amount, currency, folio FROM expenses WHERE approved = 'none' ORDER BY id DESC")
            rows = self.cursor.fetchall()
            
            approve_page.setRowCount(0)

            for row_data in rows:
                row_position = approve_page.rowCount()
                approve_page.insertRow(row_position)
                for col in range(len(row_data)):
                    approve_page.setItem(row_position, col, QTableWidgetItem(str(row_data[col])))

            # Adjust column sizes to fit content
            approve_page.resizeColumnsToContents()
        except Exception as e:
            return

    def approve_entry(self):
        selected_row = self.approve_table.currentRow()  # Use the correct table widget name
        if selected_row >= 0:
            voucher = self.approve_table.item(selected_row, 0).text()  # Use the correct table widget name

            # Update the status of the selected entry in the database
            self.cursor.execute("UPDATE expenses SET approved = 'A' WHERE voucher = ?", (voucher,))
            self.conn.commit()

            # Repopulate the "Approve Page" table to reflect the changes
            self.populate_approve_page(self.approve_table)  # Use the correct table widget name
            self.populate_table()

    def decline_entry(self):
        selected_row = self.approve_table.currentRow()  # Use the correct table widget name
        if selected_row >= 0:
            voucher = self.approve_table.item(selected_row, 0).text()  # Use the correct table widget name

            # Update the status of the selected entry in the database
            self.cursor.execute("UPDATE expenses SET approved = 'declined' WHERE voucher = ?", (voucher,))
            self.conn.commit()

            # Repopulate the "Approve Page" table to reflect the changes
            self.populate_approve_page(self.approve_table)  # Use the correct table widget name
            self.populate_table()

    
    def setup_filter_page(self):
        self.filter_tab = QWidget()
        self.tab_widget.addTab(self.filter_tab, "Filter Page")

        # Calculate the default date range
        today = QDate.currentDate()
        enddate = today.addDays(-1096)

        layout = QVBoxLayout(self.filter_tab)

        # Search bar
        self.search_bar = QLineEdit()
        self.search_bar.setPlaceholderText("Search for Description, Amount, Voucher Number or Folio...")
        layout.addWidget(self.search_bar)

        # Date range
        self.start_date_edit = QDateEdit(enddate)
        self.start_date_edit.setCalendarPopup(True)
        self.end_date_edit = QDateEdit(today)
        self.end_date_edit.setCalendarPopup(True)
        date_range_layout = QHBoxLayout()
        date_range_layout.addWidget(QLabel("Start Date:"))
        date_range_layout.addWidget(self.start_date_edit)
        date_range_layout.addWidget(QLabel("End Date:"))
        date_range_layout.addWidget(self.end_date_edit)
        layout.addLayout(date_range_layout)
    
        # Category sorting
        self.category_combo = QComboBox()
        self.category_combo.addItem("All Categories")
        self.category_combo.addItems([
            "Food & Drinks",
            "Utilities",
            "Transportation",
            "Office Supplies",
            "Entertainment",
            "Travel",
            "Education",
            "Healthcare",
            "Miscellaneous (with Folio)",
            "Charity and Donations",
            "Advertising",
            "Rent or Mortgage",
            "Repairs and Maintenance",
            "Personal Withdrawings",
            "Wages/Salaries"
        ])
        layout.addWidget(self.category_combo)

        # Currency sorting
        self.currency_combo = QComboBox()
        self.currency_combo.addItem("All Currencies")
        self.currency_combo.addItem("TZS")
        self.currency_combo.addItem("USD")
        layout.addWidget(self.currency_combo)


        # Apply Filter button
        self.apply_filter_button = QPushButton("Apply Filter")
        self.apply_filter_button.clicked.connect(self.apply_filter)
        layout.addWidget(self.apply_filter_button)

        # Add the table for displaying filtered entries
        self.filtered_table = QTableWidget()
        self.filtered_table.setColumnCount(7)  # Adjust column count based on your needs
        self.filtered_table.setHorizontalHeaderLabels(["              Voucher               ", "             Date               ", "             Description               ", "              Category                ", "                 Amount              ", " Currency ", "                 Folio                "])
            # After populating the table with data
        self.filtered_table.resizeColumnsToContents()


        # Enable sorting for the table
        self.filtered_table.setSortingEnabled(True)
        layout.addWidget(self.filtered_table)


    def apply_filter(self):
        search_text = self.search_bar.text()
        selected_category = self.category_combo.currentText()
        selected_currency = self.currency_combo.currentText()
        start_date = self.start_date_edit.date().toString("yyyy-MM-dd")
        end_date = self.end_date_edit.date().toString("yyyy-MM-dd")

        # Construct the SQL query based on filter options
        query = "SELECT voucher, date, description, category, amount, currency, folio FROM expenses WHERE (approved = 'A' OR approved = 'none')"

        params = []

        if search_text:
            query += " AND (description LIKE ? OR voucher LIKE ? OR folio LIKE ? OR amount LIKE ?)"
            params.extend([f"%{search_text}%", f"%{search_text}%", f"%{search_text}%", f"%{search_text}%"])

        if selected_category != "All Categories":
            query += " AND category = ?"
            params.append(selected_category)

        if selected_currency != "All Currencies":
            query += " AND currency = ?"
            params.append(selected_currency)

        query += " AND date BETWEEN ? AND ?"
        params.extend([start_date, end_date])

        try:
            self.cursor.execute(query, params)
            rows = self.cursor.fetchall()

            filtered_table = QTableWidget()
            
            self.filtered_table.setRowCount(0)
            for row_data in rows:
                row_position = self.filtered_table.rowCount()
                self.filtered_table.insertRow(row_position)
                for col, value in enumerate(row_data):
                    self.filtered_table.setItem(row_position, col, QTableWidgetItem(str(value)))

        except Exception as e:
            QMessageBox.critical(self, "Error", f"An error occurred: {str(e)}")

       
            
    def populate_table(self):
        base_query = "SELECT voucher, date, description, category, amount, currency, folio FROM expenses WHERE approved = 'A' OR approved = 'none'"
        order_limit_query = " ORDER BY id DESC LIMIT ?, ?"

        params = [((self.page_number - 1) * self.rows_per_page), self.rows_per_page]

        query = base_query + order_limit_query

        self.cursor.execute(query, params)
        rows = self.cursor.fetchall()

        self.table.setRowCount(0)

        for row_data in rows:
            row_position = self.table.rowCount()
            self.table.insertRow(row_position)
            for col, value in enumerate(row_data):
                self.table.setItem(row_position, col, QTableWidgetItem(str(value)))
            # Adjust column sizes to fit content
        self.table.resizeColumnsToContents()
    #correct syntax
        
    
    def next_page(self):
        self.page_number += 1
        self.populate_table()

    def prev_page(self):
        if self.page_number > 1:
            self.page_number -= 1
            self.populate_table()

    def edit_entry(self):
        selected_row = self.table.currentRow()
        if selected_row >= 0:
            voucher = self.table.item(selected_row, 0).text()

            # Fetch data for the selected voucher from the database
            self.cursor.execute("SELECT * FROM expenses WHERE voucher = ?", (voucher,))
            entry_data = self.cursor.fetchone()

            # Open a Tkinter popup window for editing
            root = tk.Tk()
            root.withdraw()

            dialog = simpledialog.Toplevel(root)
            dialog.title("Edit Expense")

            # Labels and entry widgets for editing
            tk.Label(dialog, text="Date:").pack()
            date_entry = tk.Entry(dialog)
            date_entry.insert(0, entry_data[2])  # Insert fetched date
            date_entry.pack()

            tk.Label(dialog, text="Description:").pack()
            description_entry = tk.Entry(dialog)
            description_entry.insert(0, entry_data[3])  # Insert fetched description
            description_entry.pack()

            tk.Label(dialog, text="Amount:").pack()
            amount_entry = tk.Entry(dialog)
            amount_entry.insert(0, entry_data[5])  # Insert fetched amount
            amount_entry.pack()

            tk.Label(dialog, text="Currency:").pack()
            currency_entry = tk.Entry(dialog)
            currency_entry.insert(0, str(entry_data[8]))  # Insert fetched currency
            currency_entry.pack()

            tk.Label(dialog, text="Category:").pack()

            # Create radio buttons for categories
            category_var = tk.StringVar()
            radio_buttons = []
            for idx, category in enumerate(categorieslist):
                radio_button = tk.Radiobutton(dialog, text=category, variable=category_var, value=category)
                radio_button.pack()
                radio_buttons.append(radio_button)

            # Set the selected category based on the fetched data
            fetched_category = entry_data[4]
            for idx, category in enumerate(categorieslist):
                if fetched_category == category:
                    category_var.set(category)
                    radio_buttons[idx].select()
                    break

            tk.Label(dialog, text="Folio (optional):").pack()
            folio_entry = tk.Entry(dialog)
            folio_entry.insert(0, entry_data[6])  # Insert fetched folio
            folio_entry.pack()

            def save_edited_data():
                # Get edited data from entry widgets
                edited_date = date_entry.get()
                edited_description = description_entry.get()
                edited_amount = amount_entry.get()
                edited_currency = currency_entry.get()
                edited_category = category_var.get()
                edited_folio = folio_entry.get()

                try:
                    update_query = "UPDATE expenses SET date = ?, description = ?, amount = ?, currency = ?, category = ?, folio = ?, edited = 'yes', approved = 'none' WHERE voucher = ?"
                    self.cursor.execute(update_query, (edited_date, edited_description, edited_amount, edited_currency, edited_category, edited_folio, voucher))
                    self.conn.commit()
                    print("Data updated successfully!")
                except Exception as e:
                    print(f"Error updating data: {e}")
                    self.conn.rollback()

                dialog.destroy()  # Close the dialog window
                self.populate_table()


            save_button = tk.Button(dialog, text="Save", command=save_edited_data)
            save_button.pack()
            dialog.mainloop()


    def delete_entry(self):
        selected_row = self.table.currentRow()
        if selected_row >= 0:
            voucher = self.table.item(selected_row, 0).text()
            
            # Mark the selected entry as deleted in the database
            self.cursor.execute("UPDATE expenses SET approved = 'deleted' WHERE voucher = ?", (voucher,))
            self.conn.commit()
            
            # Repopulate the table to reflect the changes
            self.populate_table()
            #populate filter page as well

    def send_email(self):
        # Criteria: Personal Withdrawings section or any amount over $1000 or TZS 1,000,000 for/of today
        criteria_query = '''SELECT voucher, date, description, category, amount, currency, folio FROM expenses 
        WHERE (approved = 'A' OR approved = 'none') 
        AND (category = 'Personal Withdrawings' 
        OR (amount > 1000 AND currency = 'USD') 
        OR (amount > 1000000 AND currency = 'TZS'))
        AND date = CURDATE()'''

        try:
            self.cursor.execute(criteria_query)
            rows = self.cursor.fetchall()

            if rows:
                # Create an email message with HTML content
                msg = MIMEMultipart()
                msg["From"] = sender_email
                msg["To"] = recipient_email
                msg["Subject"] = f"{QDate.currentDate().toString('yyyy-MM-dd')} Petty Cash Expenses Report"

                # Create an HTML table for the email body
                email_body = "<html><body>"
                email_body += "<p>The following expenses meet the criteria:</p>"
                email_body += "<table border='1' cellspacing='0' cellpadding='5'><tr><th>Voucher</th><th>Date</th><th>Description</th><th>Category</th><th>Amount</th><th>Folio</th></tr>"

                for row in rows:
                    email_body += f"<tr><td>{row[0]}</td><td>{row[1]}</td><td>{row[2]}</td><td>{row[3]}</td><td>{row[4]} {row[5]}</td><td>{row[6]}</td></tr>"

                email_body += "</table></body></html>"
                msg.attach(MIMEText(email_body, "html"))

                # Connect to the SMTP server and send the email
                server = smtplib.SMTP(smtp_server, smtp_port)
                server.starttls()
                server.login(sender_email, sender_password)
                server.sendmail(sender_email, recipient_email, msg.as_string())
                server.quit()

                QMessageBox.information(self, "Email Sent", "The email report has been sent successfully.")
            else:
                QMessageBox.information(self, "No Data", "There are no expenses that meet the criteria.")
        except Exception as e:
            QMessageBox.critical(self, "Email Error", f"An error occurred while sending the email: {str(e)}")


# Define the send_daily_email function
def send_daily_email():
    try:
        db_connection = {
    "host": host,
    "user": user,
    "password": password,
    "database": database,
}
        # Connect to the MariaDB database
        cursor = db_connection.cursor()

        # Get today's date
        today_date = QDate.currentDate().toString('yyyy-MM-dd')
        
        # Criteria: Personal Withdrawings section or any amount over $1000 or TZS 1,000,000
        criteria_query = """
            SELECT voucher, date, description, category, amount, currency, folio, approved
            FROM expenses
            WHERE date = ? 
                AND approved = 'A'
                AND (category = 'Personal Withdrawings' 
                OR (amount > 1000 AND currency = 'USD') 
                OR (amount > 1000000 AND currency = 'TZS'))
                AND date = CURDATE()
        """

        #inly today's dates stuff

        cursor.execute(criteria_query, (today_date,))
        rows = cursor.fetchall()

        if rows:
            # Create an email message
            msg = MIMEMultipart()
            msg["From"] = sender_email
            msg["To"] = recipient_email
            msg["Subject"] = f"{today_date} Petty Cash Expenses Report"       #should send in dd-MM-yyyy, or is it fine anyways/how?

            # Compose the email content in tabular form (HTML)
            email_body = """
                <html>
                <head></head>
                <body>
                <table border="1">
                <tr>
                    <th>Voucher</th>
                    <th>Date</th>
                    <th>Description</th>
                    <th>Category</th>
                    <th>Amount</th>
                    <th>Currency</th>
                    <th>Folio</th>
                    <th>Approved</th>
                </tr>
            """

            for row in rows:
                email_body += f"""
                    <tr>
                        <td>{row[0]}</td>
                        <td>{row[1]}</td>
                        <td>{row[2]}</td>
                        <td>{row[3]}</td>
                        <td>{row[4]}</td>
                        <td>{row[5]}</td>
                        <td>{row[6]}</td>
                        <td>{row[7]}</td>
                    </tr>
                """

            email_body += """
                </table>
                </body>
                </html>
            """

            msg.attach(MIMEText(email_body, "html"))

            # Connect to the SMTP server and send the email
            server = smtplib.SMTP(smtp_server, smtp_port)
            server.starttls()
            server.login(sender_email, sender_password)
            server.sendmail(sender_email, recipient_email, msg.as_string())
            server.quit()

            # Close the database connection
            cursor.close()
            db_connection.close()

        else:
            return

    except Exception as e:
        return

# Create a scheduler instance
scheduler = BackgroundScheduler()

# Schedule the email sending task to run daily at 17:00 (adjust the time as needed)
scheduler.add_job(send_daily_email, trigger='cron', hour=16, minute=45, second=0)

# Start the scheduler
scheduler.start()



if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = PettyCashApp()
    window.show()
    sys.exit(app.exec())


print("Error: " + str(Exception))